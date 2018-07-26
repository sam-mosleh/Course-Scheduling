import queue
from typing import List, Any
import multiprocessing as mp
import pandas as pd
from pandas import *
from random import *
import threading
import time
import openpyxl
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

# Defining infinity number
maxInfinity = 1000000
stageSize = 900
genomClassSize = 0

# returns random element of a list x
def randE(x, f_lastMore = 0):
    if not f_lastMore:
        ri = randint(0, len(x) - 1)
    else:
        ri = randint(3*len(x)/4, len(x)-1)
    return x[ri]


class Instructor:
    def __init__(self, f_name):
        self.instructorName = f_name
        self.courseList = []
        self.freeTimes = [0 for i in range(genomClassSize)]


class Course:
    def __init__(self, f_name, f_containing=maxInfinity, f_timesInWeek=2):
        self.courseName = f_name
        self.timesInWeek = f_timesInWeek
        self.containing = f_containing
        self.presentors = []


class Classroom:
    def __init__(self, f_name, f_capacity=maxInfinity):
        self.className = f_name
        self.capacity = f_capacity
        self.usableTimes = [0 for i in range(genomClassSize)]


class Chromosome:
    def __init__(self):
        # Size of a chromosome is 5(Days)*4(Times)*Number of classes
        # Schedule contains ID of instructor and course
        self.scheduleSize = genomClassSize * len(classrooms)
        self.schedule = [(-1, -1) for i in range(self.scheduleSize)]
        self.__myScore = 0
        self.coursePresentors = dict()
        self.distictPresentors = dict()

    def scoring(self):
        return self.__myScore

    def randomInitialize(self):
        # Fill it randomly

        for randCours in range(len(courses)):
            for repT in range(courses[randCours].timesInWeek):
                if len(courses[randCours].presentors) > 0:
                    randInst = randE(courses[randCours].presentors)
                    rs = randint(0, self.scheduleSize-1)
                    while self.schedule[rs][0] != -1:
                        rs = randint(0, self.scheduleSize - 1)
                    self.schedule[rs] = (randInst, randCours)

        self.fitnessCalculation()

    def mutate(self, f_swapProb=0.01, f_instProb=0.004, f_coursProb=0.004, f_addProb=0.01, f_inverseProb=0.001):
        randprob = random()
        if randprob <= f_swapProb:
            self.mutateBySwap()

        randprob = random()
        if randprob <= f_instProb:
            self.mutateByChangingInstructor()

        randprob = random()
        if randprob <= f_coursProb:
            self.mutateByCourse()

        randprob = random()
        if randprob <= f_addProb:
            self.mutateByAdding()

        randprob = random()
        if randprob <= f_inverseProb:
            self.mutateByInversing()

    # Change one instructor's Course
    def mutateByCourse(self):
        randSch = randrange(self.scheduleSize)
        # Is randSch valid?
        if self.schedule[randSch][0] != -1:
            randCourse = randE(instructorsList[self.schedule[randSch][0]].courseList)
            self.schedule[randSch] = self.schedule[randSch][0], randCourse

    # Change one course's Instructor
    def mutateByChangingInstructor(self):
        randSch = randrange(self.scheduleSize)
        # Is randSch valid?
        if self.schedule[randSch][0] != -1:
            randInst = randE(courses[self.schedule[randSch][1]].presentors)
            self.schedule[randSch] = randInst, self.schedule[randSch][1]

    # Add instructor to one class and time
    def mutateByAdding(self):
        randSch = randrange(self.scheduleSize)
        if self.schedule[randSch][0] == -1:
            randInst = randint(0, len(instructorsList)-1)
            if len(instructorsList[randInst].courseList)>0:
                randCours = randE(instructorsList[randInst].courseList)
                self.schedule[randSch] = randInst, randCours

    # Swap 2 blocks in chromosome doing it mutationSize times
    def mutateBySwap(self):
        randSch1 = randrange(self.scheduleSize)
        randSch2 = randrange(self.scheduleSize)
        self.schedule[randSch1], self.schedule[randSch2] = self.schedule[randSch2], self.schedule[randSch1]

    # Inverse an interval
    def mutateByInversing(self):
        randSch1 = randrange(self.scheduleSize)
        randSch2 = randrange(self.scheduleSize)
        if randSch1 > randSch2:
            randSch1, randSch2 = randSch2, randSch1
        indexNum = 0
        while randSch1 + indexNum < randSch2 - indexNum:
            self.schedule[randSch1 + indexNum], self.schedule[randSch2 - indexNum] = self.schedule[randSch2 - indexNum], self.schedule[randSch1 + indexNum]
            indexNum += 1

    # Calculate score of the chromosome
    def fitnessCalculation(self):
        score = 0
        self.coursePresentors.clear()
        self.distictPresentors.clear()
        numberOfClasses = [[0 for h in instructorsList] for k in allDays]
        instructorTeachingNoon = [0 for h in instructorsList]
        chartDayTimes = [list() for k in chartSets]
        teachersDelta = [0 for i in instructorsList]
        allDatas = 0
        for i in range(self.scheduleSize):
            if self.schedule[i][0] != -1:
                # Located with enough seats (Course cont and class cap)
                if courses[self.schedule[i][1]].containing <= classrooms[classDayTime(i)[0]].capacity:
                    score += 5
                else:
                    score -= 5

                # Professor is not busy
                _busy = 0
                _not_busy = 1
                indSch = i % genomClassSize
                flagOstad = _not_busy
                while indSch<self.scheduleSize:
                    if indSch != i and self.schedule[i][0] == self.schedule[indSch][0]:
                        flagOstad = _busy
                    indSch += genomClassSize
                if flagOstad == _not_busy:
                    score += 1
                else:
                    score -= 1

                # Course is not teaching in another class
                _not_taught = 0
                _taught = 1
                indSch = i % genomClassSize
                flagDars = _not_taught
                while indSch < self.scheduleSize:
                    if indSch != i and self.schedule[i][1] == self.schedule[indSch][1]:
                        flagDars = _taught
                    indSch += genomClassSize
                if flagDars == _not_taught:
                    score += 1
                else:
                    score -= 1

                # Instructors are available on that time
                if instructorsList[self.schedule[i][0]].freeTimes[i % genomClassSize]:
                    score += 1
                else:
                    score -= 1

                # Classes are available on that time
                if classrooms[classDayTime(i)[0]].usableTimes[i % genomClassSize]:
                    score += 1
                else:
                    score -= 1

                # Setting Course Presentors
                if self.schedule[i][1] not in self.coursePresentors:
                    self.coursePresentors[self.schedule[i][1]] = [(self.schedule[i][0], i), ]
                else:
                    self.coursePresentors[self.schedule[i][1]].append((self.schedule[i][0], i))

                # Setting Teachers Time data
                tmpSaver = classDayTime(i)
                numberOfClasses[tmpSaver[1]][self.schedule[i][0]] += 1
                if tmpSaver[2] >= 2:
                    instructorTeachingNoon[self.schedule[i][0]] = 1

                # Setting Chart Data
                for setIndex in range(len(chartSets)):
                    if self.schedule[i][1] in chartSets[setIndex]:
                        tmpSaver = classDayTime(i)
                        chartDayTimes[setIndex].append((tmpSaver[1], tmpSaver[2]))

                # Setting instructors Data for Variance
                teachersDelta[self.schedule[i][0]] += 1
                allDatas += 1

        for i in self.coursePresentors:
            # Taught too many times
            score += 10
            if len(self.coursePresentors[i]) > courses[i].timesInWeek:
                score -= (len(self.coursePresentors[i]) - courses[i].timesInWeek) * 10
            # Taught more than one teacher
            score += 10
            self.distictPresentors[i] = len(set([x for x, y in self.coursePresentors[i]]))
            score -= self.distictPresentors[i] * 10

        # Instructors are on morning
        for x in instructorTeachingNoon:
            if not x:
                score += 5
        # Instructors having less than 3 courses on a day
        for x in range(len(allDays)):
            for y in range(len(instructorsList)):
                if numberOfClasses[x][y] <= 3:
                    score += 1

        # Chart courses dont match day and time
        for x in chartDayTimes:
            if len(x) == len(set(x)):
                score += 1

        # Setting Variance
        dataMean = allDatas / len(instructorsList)
        dataVariance = 0
        for d in teachersDelta:
            dataVariance += (d - dataMean)*(d - dataMean)
        score -= 1 * int(dataVariance)
        self.__myScore = score


def classDayTime(f_n):
    classN = f_n // genomClassSize
    f_n %= genomClassSize
    dayN = f_n // 5
    timeN = f_n % 5
    return classN, dayN, timeN


def readFromExcel(f_profSkill, f_profTime, f_freeClass, f_courseRegister, f_classCap):
    global allDays, classrooms, instructorsList, courses, universityTimes, genomClassSize
    # Opening Excels
    classRead = pd.ExcelFile(f_freeClass)
    profskillRead = pd.read_excel(f_profSkill)
    allDayRead = pd.read_excel(f_freeClass, 0)
    registerRead = pd.read_excel(f_courseRegister)
    capRead = pd.read_excel(f_classCap)
    chartRead = pd.read_excel("Chart.xlsx")

    # Import Course Times from FreeClass table
    for ut in allDayRead.columns:
        universityTimes.append(ut)
    print(universityTimes)

    # Import Days from FreeClass table
    for d in allDayRead.index:
        allDays.append(d)
    print(allDays)

    # Set size of a class data
    genomClassSize = len(allDays) * len(universityTimes)

    # Import classes from FreeClass table and then import cap from class cap table
    classMap = dict()
    for selectedClass in classRead.sheet_names:
        readSheet = pd.read_excel(f_freeClass, sheet_name=selectedClass)
        newClass = Classroom(selectedClass)
        for k in range(len(readSheet.index)):
            for j in range(len(readSheet.columns)):
                if readSheet[readSheet.columns[j]][readSheet.index[k]]:
                    newClass.usableTimes[len(universityTimes) * k + j] = 1
        classrooms.append(newClass)

    for cIndex in range(len(classrooms)):
        classMap[classrooms[cIndex].className] = cIndex
    print([c.className for c in classrooms])

    for k in capRead.index:
        if capRead[capRead.columns[1]][k]:
            classrooms[classMap[str(capRead[capRead.columns[0]][k])]].capacity = 15

    # Import Courses from ProfSkill table and then import cap from register table
    for selectedCourse in profskillRead.columns:
        tmpData = selectedCourse.split("-")
        #print(tmpData)
        newCourse = Course(tmpData[0])
        if tmpData[1] == "1" or tmpData[1] == "2":
            newCourse.timesInWeek = 1
        courses.append(newCourse)
        #print(newCourse.courseName, newCourse.timesInWeek)

    courseMap = dict()
    for cIndex in range(len(courses)):
        courseMap[courses[cIndex].courseName] = cIndex
    print([c.courseName for c in courses])
    for k in registerRead.index:
        if registerRead[registerRead.columns[1]][k]:
            tmpData = registerRead[registerRead.columns[0]][k].split("-")
            #print("here to set>", tmpData)
            courses[courseMap[tmpData[0]]].containing = 15
    print("number of all courses is =>", len(courses))

    # Import Instructors
    for i in profskillRead.index:

        # Name of Instructor
        newIns = Instructor(i)

        # Import Time of Instructor from ProfFreeTime
        prTimeRead = pd.read_excel(f_profTime, sheet_name=i)
        for j in range(len(prTimeRead.columns)):
            for k in range(len(prTimeRead.index)):
                if prTimeRead[prTimeRead.columns[j]][prTimeRead.index[k]]:
                    newIns.freeTimes[len(universityTimes) * k + j] = 1

        # Courses of Instructor
        for j in range(len(profskillRead.columns)):
            if profskillRead[profskillRead.columns[j]][i]:
                newIns.courseList.append(j)
                courses[j].presentors.append(len(instructorsList))
        instructorsList.append(newIns)

    for chartCol in chartRead.columns:
        chartSets.append(set())
        for charInd in chartRead.index:
            if type(chartRead[chartCol][charInd]) != float:
                chartSets[-1].add((chartRead[chartCol][charInd].split("-"))[0])


def initChromosomes(f_numberOfNodes, f_chromoList):
    for i in range(f_numberOfNodes):
        newChro = Chromosome()
        newChro.randomInitialize()
        f_chromoList.append(newChro)


# Offsprings of Crossover of Two Chromosomes
def crossover(chrom1: Chromosome, chrom2: Chromosome, f_crossoverProb, f_crossoverPointNumber=8):
    firstChromo = Chromosome()
    secondChromo = Chromosome()
    if f_crossoverProb < randrange(100):
        crossoverPoints = {}
        while len(crossoverPoints) < f_crossoverPointNumber:
            randPoint = randrange(stageSize)
            if randPoint not in crossoverPoints:
                crossoverPoints[randPoint] = 1
        firstChoice = randint(0, 1)
        for i in range(chrom1.scheduleSize):
            if firstChoice:
                firstChromo.schedule[i] = chrom1.schedule[i]
                secondChromo.schedule[i] = chrom2.schedule[i]
            else:
                firstChromo.schedule[i] = chrom2.schedule[i]
                secondChromo.schedule[i] = chrom1.schedule[i]
            if i in crossoverPoints:
                firstChoice = (not firstChoice)
        firstChromo.fitnessCalculation()
        secondChromo.fitnessCalculation()
    else:
        firstChromo = chrom1
        secondChromo = chrom2
    return firstChromo, secondChromo


def writeResults(f_tableName):
    global universityTimes
    teachers = [["-" for i in range(genomClassSize)] for j in instructorsList]
    mytxt = open(f_tableName + ".txt", "w")
    coursesOfInstructor = [[] for i in instructorsList]
    for i in range(resultChro.scheduleSize):
        if resultChro.schedule[i][0] != -1:
            coursesOfInstructor[resultChro.schedule[i][0]].append(resultChro.schedule[i][1])
            # mytxt.write(instructorsList[resultChro.schedule[i][0]].instructorName +
            #             " - " + courses[resultChro.schedule[i][1]].courseName+"\n")
            teachers[resultChro.schedule[i][0]][i % genomClassSize] = courses[resultChro.schedule[i][1]].courseName + \
                                                                      " - " + \
                                                                   classrooms[classDayTime(i)[0]].className

    for i in range(len(coursesOfInstructor)):
        mytxt.write(instructorsList[i].instructorName + " : ")
        for j in coursesOfInstructor[i]:
            mytxt.write(courses[j].courseName + " ")
        mytxt.write("\n")
    mytxt.write(str(resultChro.scoring()))
    mytxt.close()
    myWorkBook = openpyxl.Workbook()
    newSheet = myWorkBook.active

    newSheet.title = instructorsList[0].instructorName
    for c in range(len(universityTimes)):
        newSheet.cell(1, c+2).value = universityTimes[c]
    for r in range(len(allDays)):
        newSheet.cell(r+2, 1).value = allDays[r]
    for j in range(genomClassSize):
        newSheet.cell(classDayTime(j)[1] + 2, classDayTime(j)[2] + 2).value = teachers[0][j]

    for i in range(len(instructorsList) - 1):
        myWorkBook.create_sheet(title=instructorsList[i+1].instructorName)
        newSheet = myWorkBook[instructorsList[i+1].instructorName]
        for c in range(len(universityTimes)):
            newSheet.cell(1, c + 2).value = universityTimes[c]
        for r in range(len(allDays)):
            newSheet.cell(r + 2, 1).value = allDays[r]
        for j in range(genomClassSize):
            newSheet.cell(classDayTime(j)[1] + 2, classDayTime(j)[2] + 2).value = teachers[i + 1][j]
    myWorkBook.save(f_tableName + ".xlsx")


# Gets sorted chromList and select by rank of chromosome
def selectRandomByRank(f_chromList: List[Chromosome]):
    listS = f_chromList.__len__()
    uniformRandomSelect = randrange(listS*(listS + 1) // 2)  # Sum of all ranks
    #print("my random:",uniformRandomSelect)
    # find the selection by binary search ==> [right,left]
    rightBound = listS - 1
    leftBound = 0
    mid = 0
    while rightBound > leftBound:
        mid = (rightBound + leftBound) // 2
        nowSum = ((mid + 1) * (mid + 2)) // 2
        #print(leftBound, mid, rightBound, nowSum, uniformRandomSelect)#, f_chromList[leftBound].scoring(), f_chromList[mid].scoring(), f_chromList[rightBound].scoring(),)
        if uniformRandomSelect < nowSum:
            rightBound = mid
        elif uniformRandomSelect > nowSum:
            leftBound = mid + 1
        else:
            break
    #print(f_chromList[leftBound].scoring())
    return f_chromList[mid]


def selectRandomByRWS(f_chromList: List[Chromosome]):
    tmpSum = 0
    for i in f_chromList:
        tmpSum += i.scoring()
    uniformRandomSelect = randrange(tmpSum)  # Sum of all ranks
    resultInd = 0
    while uniformRandomSelect > 0:
        uniformRandomSelect -= f_chromList[resultInd].scoring()
        resultInd += 1
    return f_chromList[resultInd-1]


# X for main pop with negativity 0 and search pop with negativity 1
def crossoverByCorrolate(f_chromeA, f_chromeB, f_negativity):
    dist = 0
    for i in range(f_chromeA.scheduleSize):
        if f_chromeA.schedule[i] != f_chromeB.schedule[i]:
            dist += 1
    s = dist/stageSize
    if f_negativity:
        if s > 0.8:
            return crossover(f_chromeA, f_chromeB, 20)
        elif s < 0.2:
            return crossover(f_chromeA, f_chromeB, 100)
        else:
            return crossover(f_chromeA, f_chromeB, 100 - 100 * s)
    else:
        if s > 0.8:
            return crossover(f_chromeA, f_chromeB, 100)
        elif s < 0.2:
            return crossover(f_chromeA, f_chromeB, 20)
        else:
            return crossover(f_chromeA, f_chromeB, 100 * s)


def makeGeneration(f_nowGeneration: List[Chromosome], f_ExploitOrExplore):
    nextGeneration = [] # type: List[Chromosome]
    for popIteration in range(stageSize//4):
        firstOffspring = Chromosome()
        secondOffspring = Chromosome()
        if f_ExploitOrExplore == 0:
            # Exploit
            firstParentChromo = selectRandomByRank(f_nowGeneration)
            secondParentChromo = selectRandomByRank(f_nowGeneration)
            firstOffspring, secondOffspring = crossoverByCorrolate(
                firstParentChromo, secondParentChromo, f_ExploitOrExplore)
            firstOffspring.mutate()
            secondOffspring.mutate()
        else:
            # Explore
            firstParentChromo = selectRandomByRWS(f_nowGeneration)
            secondParentChromo = selectRandomByRWS(f_nowGeneration)
            firstOffspring, secondOffspring = crossoverByCorrolate(firstParentChromo, secondParentChromo,
                                                                   f_ExploitOrExplore)
            firstOffspring.mutate(0.05, 0.02, 0.02, 0.05, 0.005)
            secondOffspring.mutate(0.05, 0.02, 0.02, 0.05, 0.005)
        firstOffspring.fitnessCalculation()
        secondOffspring.fitnessCalculation()
        # #tmpSaver = firstOffspring.scoring()
        # firstOffspring = repairChromosomeByDeleting(firstOffspring)
        # #print("after:", tmpSaver, firstOffspring.scoring())
        # #print("first offspring")
        # secondOffspring = repairChromosomeByDeleting(secondOffspring)
        # #print("second offspring")

        nextGeneration.append(firstOffspring)
        nextGeneration.append(secondOffspring)
    return nextGeneration


def iterateSemiGen(f_gen: List[Chromosome], f_ExploitOrExplore, semiGenNum=1):
    nowGen = f_gen[:]
    for i in range(semiGenNum):
        #print(i)
        nowGen = makeGeneration(nowGen, f_ExploitOrExplore)
    return nowGen


def addBinarySearch(f_list, f_elem):
    leftBound = 0
    rightBound = len(f_list) - 1
    mid = 0
    theScore = f_elem.scoring()
    while rightBound >= leftBound:
        mid = (leftBound + rightBound) // 2
        if f_list[mid].scoring() > theScore:
            rightBound = mid - 1
        elif f_list[mid].scoring() < theScore:
            leftBound = mid + 1
        else:
            break
    if f_list[mid].scoring() == theScore:
        f_list.insert(mid, f_elem)
    else:
        f_list.insert(leftBound, f_elem)


def populationsMigration(f_popA, f_popB):
    firstPopIndex = len(f_popA) - 1
    secondPopIndex = len(f_popB) - 1

    # Biggest in B goes to A
    if f_popA[-1].scoring() < f_popB[-1].scoring():
        while f_popA[-1].scoring() < f_popB[secondPopIndex].scoring():
            secondPopIndex -= 1
        secondPopIndex += 1

        changeSize = len(f_popB) - secondPopIndex
        f_popA.extend(f_popB[secondPopIndex:])
        del f_popA[stageSize//2:stageSize//2 + changeSize]
    # Biggest in A goes to B
    elif f_popA[-1].scoring() > f_popB[-1].scoring():
        while f_popB[-1].scoring() < f_popA[firstPopIndex].scoring():
            firstPopIndex -= 1
        firstPopIndex += 1

        changeSize = len(f_popA) - firstPopIndex
        f_popB.extend(f_popA[firstPopIndex:])
        del f_popA[stageSize // 2:stageSize // 2 + changeSize]

    firstPopIndex = 0
    secondPopIndex = 0
    # Smallest in B goes to A
    if f_popA[0].scoring() > f_popB[0].scoring():
            while f_popA[0].scoring() > f_popB[secondPopIndex].scoring():
                secondPopIndex += 1

            del f_popA[stageSize // 2:stageSize // 2 + secondPopIndex]
            f_popA = f_popB[ : secondPopIndex] + f_popA
    # Smallest in A goes to B
    elif f_popB[0].scoring() > f_popA[0].scoring():
            while f_popB[0].scoring() > f_popA[firstPopIndex].scoring():
                firstPopIndex += 1

            del f_popB[stageSize // 2 : stageSize // 2 + firstPopIndex]
            f_popB = f_popA[:firstPopIndex] + f_popB


# Tabu search for deleting
def repairChromosomeByDeleting(f_chromo, f_maxNumberOfSearching=1):
    # Tabu search
    searchedNum = 0
    #tabuList = set()
    #tabuList.add(f_chromo)
    bestofAll = f_chromo  # type: Chromosome
    bestSearchCandidate = f_chromo

    # coursePresentors = {}
    # for cIndex in range(bestofAll.scheduleSize):
    #     if bestofAll.schedule[cIndex][0] != -1:
    #         if bestofAll.schedule[cIndex][1] not in coursePresentors:
    #             coursePresentors[bestofAll.schedule[cIndex][1]] = [(bestofAll.schedule[cIndex][0], cIndex), ]
    #         else:
    #             coursePresentors[bestofAll.schedule[cIndex][1]].append((f_chromo.schedule[cIndex][0], cIndex))
    #print(coursePresentors)
    #print("best of all=>", bestofAll.scoring())
    # For every course search the number of presentors and distinct presentors
    chromosomeAfterDelete = Chromosome()
    chromosomeAfterDelete.schedule = bestofAll.schedule[:]
    for selectedCourse in bestofAll.coursePresentors:
        if bestofAll.coursePresentors[selectedCourse].__len__() > courses[selectedCourse].timesInWeek or\
                bestofAll.distictPresentors[selectedCourse] > 1:
            maxScore = -1
            #print(selectedCourse, changeFlag)
            for teacherName, selectIndex in bestofAll.coursePresentors[selectedCourse]:

                chromosomeAfterDelete.schedule[selectIndex] = (-1, -1)
                chromosomeAfterDelete.fitnessCalculation()
                #print("Count for fitness!", chromosomeAfterDelete.scoring(), selectedCourse)
                if maxScore < chromosomeAfterDelete.scoring():# and chromosomeAfterDelete not in tabuList:
                    maxScore = chromosomeAfterDelete.scoring()
                    bestSearchCandidate = chromosomeAfterDelete
                chromosomeAfterDelete.schedule[selectIndex] = (teacherName, selectedCourse)
            #bestofAll = bestSearchCandidate.copy()
            bestofAll = bestSearchCandidate
            #tabuList.add(bestSearchCandidate)
            searchedNum += 1
            if searchedNum >= f_maxNumberOfSearching:
                break
    #print("shitty parts =", searchedNum)
    #f_chromo = bestofAll
    return bestofAll


# Dual Population Genetic Algorithm
def dualPopProcess(f_numberOfIterations: int, f_numberOfElitism: int):
    lastGenerationPop = []  # type: List[Chromosome]
    mainPop = []  # type: List[Chromosome]
    searchPop = []  # type: List[Chromosome]
    initChromosomes(stageSize, lastGenerationPop)
    # for x in lastGenerationPop:
    #     print("first:", x)
    #     repairChromosomeByDeleting(x)
    lastGenerationPop.sort(key=lambda x: x.scoring())

    lastGenerationPop2 = []  # type: List[Chromosome]
    mainPop2 = []  # type: List[Chromosome]
    searchPop2 = []  # type: List[Chromosome]
    initChromosomes(stageSize, lastGenerationPop2)
    # for x in lastGenerationPop2:
    #     print("second:", x)
    #     repairChromosomeByDeleting(x)
    lastGenerationPop2.sort(key=lambda x: x.scoring())

    for iterNum in range(f_numberOfIterations):
        tmpTime = time.time()
        newGenerationPop = []  # type: List[Chromosome]
        newGenerationPop2 = []  # type: List[Chromosome]
        print(iterNum, ":", lastGenerationPop[-1].scoring(), lastGenerationPop2[-1].scoring())
        if iterNum == 30:
            for h in lastGenerationPop:
                print(h.scoring())
        myPool = mp.Pool()
        processList = list()
        processList.append(myPool.apply_async(iterateSemiGen, (lastGenerationPop, 0)))  # Do main population
        processList.append(myPool.apply_async(iterateSemiGen, (lastGenerationPop, 1)))  # Do searching pop
        processList.append(myPool.apply_async(iterateSemiGen, (lastGenerationPop2, 0)))
        processList.append(myPool.apply_async(iterateSemiGen, (lastGenerationPop2, 1)))
        myPool.close()
        myPool.join()

        mainPop = processList[0].get()[:]
        searchPop = processList[1].get()[:]
        mainPop2 = processList[2].get()[:]
        searchPop2 = processList[3].get()[:]

        newGenerationPop.extend(mainPop)
        newGenerationPop.extend(searchPop)
        newGenerationPop.sort(key=lambda x: x.scoring())

        newGenerationPop2.extend(mainPop2)
        newGenerationPop2.extend(searchPop2)
        newGenerationPop2.sort(key=lambda x: x.scoring())

        # Elitism
        if f_numberOfElitism > 0:
            newGenerationPop = newGenerationPop[f_numberOfElitism:]
            for bigLast in lastGenerationPop[(-1) * f_numberOfElitism:]:
                addBinarySearch(newGenerationPop, bigLast)
        lastGenerationPop = newGenerationPop[:]

        if f_numberOfElitism > 0:
            newGenerationPop2 = newGenerationPop2[f_numberOfElitism:]
            for bigLast in lastGenerationPop2[(-1) * f_numberOfElitism:]:
                addBinarySearch(newGenerationPop2, bigLast)
        lastGenerationPop2 = newGenerationPop2[:]

        populationsMigration(lastGenerationPop, lastGenerationPop2)
        print("one loop time is =>", time.time() - tmpTime)

    return lastGenerationPop, lastGenerationPop2


# Initialize Global variables

allDays = []
universityTimes = []
classrooms = []  # type: List[Classroom]
courses = []  # type: List[Course]
instructorsList = []  # type: List[Instructor]
#chromosomeList = []  # type: List[Chromosome]
resultChro = Chromosome()
chartSets = []  # type: List[set]
if __name__ == "__main__":
    times = [("0", "10"), ("1", "20")]
    skill = [("0", "10"), ("6", "40")]
    freeclasses = ["1", ]
    classcaps = ["1", ]
    coursereg = ["1", ]
    for x, y in times:
        for a, b in skill:
            for fc in freeclasses:
                for cr in coursereg:
                    for cp in classcaps:
                        if b == y:
                            print(x, y, a, b, fc, cr, cp)

                            # Clear For New Data
                            universityTimes.clear()
                            courses.clear()
                            instructorsList.clear()
                            allDays.clear()
                            classrooms.clear()
                            resultChro = Chromosome()
                            # Start reading

                            readFromExcel("profskill" + a + "_profnumber-" + b + ".xlsx",
                                          "prof_freetime" + a + "_profnumber-" + b + ".xlsx",
                                          "Freeclass" + fc + ".xlsx",
                                          "register" + cr + ".xlsx",
                                          "class_capacity" + cp + ".xlsx")

                            # Initialize the Stage list0
                            #initChromosomes(stageSize)
                            # Start timing and processing
                            beforeStarting = time.time()
                            # Run new processes on 2 Threads! then make it 4
                            myTmp = dualPopProcess(300, 5)
                            if myTmp[0][-1].scoring() > myTmp[1][-1].scoring():
                                resultChro = myTmp[0][-1]
                            else:
                                resultChro = myTmp[1][-1]
                            print(x, y, a, b, " done!:", resultChro.scoring())
                            print(time.time() - beforeStarting)
                            # Write data
                            writeResults("result_"+a+"_"+x+"_"+fc+"_"+cp+"_"+cr)
